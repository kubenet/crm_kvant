from pathlib import Path
import openpyxl, openpyxl.utils
import datetime
from docxtpl import DocxTemplate
import pymorphy2

group = 'IT_project.xlsx'
student = 'Золотарев Максим Василевьич'
nametag = ['Surn', 'Name', 'Patr']
morph = pymorphy2.MorphAnalyzer()

lists_path = Path('lists')
students = []


def xlStudFiller(group):
    wb = openpyxl.load_workbook(lists_path / group)
    students = wb.sheetnames
    return students


def xlStudProd(group, student):
    wb = openpyxl.load_workbook(lists_path / group)

    ws = wb['GROUP_INFO']
    SECTION = ws['B1'].value
    GROUP_NAME = ws['B2'].value
    DATE_START = ws['B10'].value
    DATE_END = ws['B11'].value
    HOURS = ws['B12'].value

    ws = wb[student]

    SEX = ws['B17'].value
    STUDENT_NAME = ''
    student = student.split()
    for i, name in enumerate(student):
        for j, p in enumerate(morph.parse(name)):
            if nametag[i] in p.tag and SEX in p.tag:
                STUDENT_NAME += morph.parse(name)[j].inflect({'accs'}).word + ' '

    TODAY = datetime.datetime.today().strftime("%d.%m.%Y")
    FIO_Student = STUDENT_NAME.title()
    BIRTHDATE = ws['B4'].value
    YEARS = ws['C4'].value
    SCHOOL = ws['B5'].value
    CLASS = ws['B7'].value
    FIO_Parent = ws['B12'].value
    PAR_Serial = ws['B14'].value
    PAR_Num = ws['C14'].value
    PAR_PASS_PLACE = ws['B15'].value
    LIVING_ADRESS = ws['B16'].value

    doc = DocxTemplate('zayavlenie_do_14_let_novaya_forma.docx')
    context =  {'SECTION': SECTION,
                'GROUP_NAME': GROUP_NAME,
                'DATE_START': DATE_START,
                'DATE_END': DATE_END,
                'HOURS': HOURS,
                'TODAY': TODAY,
                'FIO_Student': FIO_Student,
                'BIRTHDATE': BIRTHDATE,
                'YEARS': YEARS,
                'SCHOOL': SCHOOL,
                'CLASS': CLASS,
                'FIO_Parent': FIO_Parent,
                'PAR_Serial': PAR_Serial,
                'PAR_Num': PAR_Num,
                'PAR_PASS_PLACE': PAR_PASS_PLACE,
                'LIVING_ADRESS': LIVING_ADRESS}
    doc.render(context)
    doc.save("zayavlenie.docx")
    print(BIRTHDATE)


xlStudProd(group, student)
